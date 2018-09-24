#! python3
# spreadsheet_cell_inverter.py
# usage: spreadsheet_cell_inverter.py {source_xlsx} {output_xlsx}

import sys
import openpyxl

# get source file and output file names
if len(sys.argv) > 2:
    source_file = sys.argv[1]
    output_file = sys.argv[2]

    # load active sheets for both
    source_wb = openpyxl.load_workbook(source_file)
    source_sheet = source_wb.active
    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active

    # copy from source sheet to output sheet, inverting cols and rows
    for row in range(1, source_sheet.max_row + 1):
        for col in range(1, source_sheet.max_column + 1):
            output_sheet.cell(row=col, column=row).value = source_sheet.cell(row=row, column=col).value

    # save
    output_wb.save(output_file)

else:
    print('usage: spreadsheet_cell_inverter.py {source_xlsx} {output_xlsx}')
