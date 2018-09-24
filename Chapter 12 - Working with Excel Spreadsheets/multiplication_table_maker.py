#! python3
# multiplication_table_maker.py
# usage: python multiplication_table_maker.py {num}

import sys
import openpyxl
from openpyxl.styles import Font

# get the number to make the table for
if len(sys.argv) > 1:
    num = int(sys.argv[1])
    # open new workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    # init the row and column numbers
    for i in range(2, num + 2):
        sheet.cell(row=1, column=i).value = i - 1
        sheet.cell(row=i, column=1).value = i - 1
    # get row 2 and col 2 numbers, multiply them together
    for row in range(2, sheet.max_row + 1):
        for col in range(2, sheet.max_column + 1):
            row_value = sheet.cell(row=1, column=col).value
            col_value = sheet.cell(row=row, column=1).value
            sheet.cell(row=row, column=col).value = row_value * col_value
    # bold the rows and columns
    for cell in sheet['A']:
        cell.font = Font(b=True)
    for cell in sheet['1']:
        cell.font = Font(b=True)
    # save table
    wb.save(f'multiplication table {num}.xlsx')
else:
    # if no number provided, print usage
    print("usage: python multiplication_table_maker.py {num}")
