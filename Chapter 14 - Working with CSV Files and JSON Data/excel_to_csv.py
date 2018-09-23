#! python3
# excel_to_csv.py
# usage: excel_to_csv.py
# converts excel files in current working directory to csv files

import openpyxl
import csv
import os

# locate all excel files in directory
for file_ends_with_xlsx in os.listdir('.'):
    if file_ends_with_xlsx.endswith('.xlsx'):

        # go to first excel file and load it
        xlsv_openpyxl_file = openpyxl.load_workbook(file_ends_with_xlsx)

        # open sheet and new csv
        for sheet in xlsv_openpyxl_file.sheetnames:
            active_sheet = xlsv_openpyxl_file[sheet]

            # save open sheet to csv
            csv_file_name = f'{file_ends_with_xlsx[:-5]}_{sheet}.csv'
            with open(csv_file_name, 'w', newline='') as csv_file_obj:
                csv_file = csv.writer(csv_file_obj)

                # copy each row to list and write list to csv
                for row in active_sheet.iter_rows():
                    xlsx_row = []
                    for cell in row:
                        xlsx_row.append(cell.value)
                    csv_file.writerow(xlsx_row)

        xlsv_openpyxl_file.close()
