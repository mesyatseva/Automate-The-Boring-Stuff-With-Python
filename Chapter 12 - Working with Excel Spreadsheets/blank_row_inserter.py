#! python3
# blank_row_inserter.py
# usage: python blank_row_inserter.py {row} {num_of_rows} {file_name}

import openpyxl
import sys

if len(sys.argv) > 3:
    # which row to insert
    row = int(sys.argv[1])
    # num of rows to insert
    num_of_rows = int(sys.argv[2])
    # file name
    file_name = sys.argv[3]

    # load workbook and select sheet
    source_wb = openpyxl.load_workbook(file_name)
    output_wb = openpyxl.Workbook()
    source_sheet = source_wb.active
    output_sheet = output_wb.active

    # loop through until reach last row before gap
    for row in range(1, row + 1):
        for col in range(1, source_sheet.max_column + 1):
            output_sheet.cell(row=row, column=col).value = source_sheet.cell(row=row, column=col).value

    # loop including gap to end, inserts gaps
    for row in range(row + 1, source_sheet.max_row + 1):
        for col in range(1, source_sheet.max_column + 1):
            output_sheet.cell(row=row + num_of_rows, column=col).value = source_sheet.cell(row=row, column=col).value
    # save output file
    output_wb.save(f'updated_{file_name}')
else:
    print("usage: python blank_row_inserter.py {row} {num_of_rows} {file_name}")
