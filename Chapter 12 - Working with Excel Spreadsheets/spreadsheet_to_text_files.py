#! python3
# spreadsheet_to_text_files.py
# usage: python spreadsheet_to_text_files.py {input_filename.xlsx} {txt_file_name_num.txt}

import sys
import openpyxl

# loads input spreadsheet and base file name for text file
if len(sys.argv) > 2:
    input_filename = sys.argv[1]
    txt_file_name = sys.argv[2]
    # get active sheet
    wb = openpyxl.load_workbook(input_filename)
    sheet = wb.active

    # iterate through rows for each column, with each row writen to a new text doc
    count = 1
    for col in range(1, sheet.max_column + 1):
        with open(f'{count}_' + txt_file_name, 'w') as file:
            for row in sheet.iter_rows(min_col=col, max_col=col):
                file.write(str(row[0].value) + '\n')
        count += 1
